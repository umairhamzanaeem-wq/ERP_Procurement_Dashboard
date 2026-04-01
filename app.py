import io
import json
import logging
import os
import re
import unicodedata
from difflib import SequenceMatcher
from typing import Any, Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

logger = logging.getLogger(__name__)

C_SOURCE = "Source"
C_DATE = "Date"
C_PERIOD = "Period"
C_SUPPLIER_ID = "SupplierID"
C_SUPPLIER_NAME = "SupplierName"
C_DEBIT = "Debit"
C_CREDIT = "Credit"
C_AMOUNT = "Amount"
C_DESCRIPTION = "Description"
C_GL_ACCOUNT = "GLAccount"
C_COST_CENTER = "CostCenter"
C_COMPANY = "Company"
C_CATEGORY = "Category"
C_SHEET = "Sheet"

_COLUMN_SIGNATURES: dict[str, list[str]] = {
    C_DATE: [
        "datum", "date", "leistungsdatum", "buchungsdatum", "belegdatum",
        "rechnungsdatum", "invoice date", "posting date", "erfassungsdatum",
        "valutadatum", "due date", "wertstellung",
    ],
    C_PERIOD: [
        "periode", "period", "monat", "month", "zeitraum", "quartal",
        "quarter", "jahr", "year",
    ],
    C_SUPPLIER_ID: [
        "lieferant nummer", "lieferantnummer", "lieferantennummer",
        "kreditoren nr", "kreditorennr", "kreditoren nummer", "kreditorennummer",
        "supplier id", "supplierid", "konto", "lieferant nummer 2",
        "dim_subledgeraccount", "kontonummer", "account number", "account no",
        "debitor nr", "debitornr", "kunden nr", "kundennr", "vendor id",
        "vendor no", "vendor number", "lieferanten nummer", "manufacturer code",
        "lieferant code", "lieferantcode", "creditor code", "creditorcode",
    ],
    C_SUPPLIER_NAME: [
        "lieferant name", "lieferantname", "lieferant", "kreditorenname",
        "supplier name", "suppliername", "beschriftung", "kurzbezeichnung",
        "name", "firma", "vendor name", "lieferantenname", "kontobezeichnung",
        "bezeichnung", "account name", "manufacturer name", "manufacturedname",
        "creditor name", "creditorname",
    ],
    C_DEBIT: [
        "soll", "debit", "umsatz soll", "soll betrag", "sollbetrag",
        "debit amount",
    ],
    C_CREDIT: [
        "haben", "credit", "umsatz haben", "haben betrag", "habenbetrag",
        "credit amount",
    ],
    C_AMOUNT: [
        "summe", "amount", "betrag", "netto", "net", "gesamtbetrag",
        "total", "wert", "value", "brutto", "gross", "rechnungsbetrag",
        "invoice amount", "umsatz", "turnover", "nettobetrag", "bruttobetrag",
        "endbetrag", "gesamt", "amt", "total amount", "amount total",
        "invoice total", "line total", "net amount", "gross amount",
    ],
    C_DESCRIPTION: [
        "beschreibung", "description", "text", "buchungstext", "bemerkung",
        "kommentar", "comment", "note", "anmerkung", "verwendungszweck",
        "purpose", "referenz", "reference", "betreff", "subject",
        "belegtext", "posting text",
    ],
    C_GL_ACCOUNT: [
        "sachkonto", "glaccount", "gl account", "dim_glaccount", "gegenkonto",
        "hauptbuch", "ledger account", "kontenrahmen", "sachkontonummer",
        "account", "konto (gl)",
    ],
    C_COST_CENTER: [
        "kostenstelle", "cost center", "dim_costcenter", "kost1", "kost2",
        "profit center", "profitcenter",
    ],
    C_CATEGORY: [
        "kategorie", "category", "warengruppe", "product group", "gruppe",
        "group", "materialgruppe", "material group", "typ", "type",
        "art", "kind", "klasse", "class",
    ],
}

_TYPE_DESCRIPTOR_KEYWORDS = frozenset({
    "varchar", "int", "double", "string", "float", "text",
    "ganzzahl (int)", "ganzzahl", "nvarchar", "datetime", "decimal",
})

_MONTH_NAMES_DE = {
    "januar": 1, "jan": 1, "februar": 2, "feb": 2, "maerz": 3, "mar": 3,
    "april": 4, "apr": 4, "mai": 5, "juni": 6, "jun": 6,
    "juli": 7, "jul": 7, "august": 8, "aug": 8, "september": 9, "sep": 9,
    "oktober": 10, "okt": 10, "november": 11, "nov": 11,
    "dezember": 12, "dez": 12,
}

_REFERENCE_EXTRA: dict[str, list[str]] = {}
_REFERENCE_LOADED = False

_XLSX_ROW_WARN = 200_000

# Merge supplier labels when fuzzy similarity ≥ this (88% = 0.88); debit/credit then sum per merged name.
FUZZY_SUPPLIER_MERGE_THRESHOLD = 0.95


def _load_reference_schema_aliases() -> None:
    """Optional reference_schema.json next to app.py: { \"Amount\": [\"Amt\", ...], ... }."""
    global _REFERENCE_LOADED, _REFERENCE_EXTRA
    if _REFERENCE_LOADED:
        return
    _REFERENCE_LOADED = True
    base = os.path.dirname(os.path.abspath(__file__))
    for fname in ("reference_schema.json", "REFERENCE_SCHEMA.json"):
        path = os.path.join(base, fname)
        if not os.path.isfile(path):
            continue
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                continue
            for key, vals in data.items():
                if key not in _COLUMN_SIGNATURES or not isinstance(vals, (list, tuple)):
                    continue
                _REFERENCE_EXTRA[key] = [str(v).strip().lower() for v in vals if str(v).strip()]
            logger.info("Loaded reference column aliases from %s", fname)
            break
        except Exception as e:
            logger.warning("Could not load %s: %s", path, e)


def _patterns_for_column(canonical: str) -> list[str]:
    _load_reference_schema_aliases()
    base = list(_COLUMN_SIGNATURES.get(canonical, []))
    base.extend(_REFERENCE_EXTRA.get(canonical, []))
    return base


def _best_column_match(raw_col: str) -> Optional[str]:
    normed = raw_col.strip().lower().replace("_", " ").replace("-", " ")
    for canonical in _COLUMN_SIGNATURES:
        patterns = _patterns_for_column(canonical)
        if normed in patterns:
            return canonical
    for canonical in _COLUMN_SIGNATURES:
        patterns = _patterns_for_column(canonical)
        for pat in patterns:
            if pat in normed or normed in pat:
                return canonical
    for canonical in _COLUMN_SIGNATURES:
        patterns = _patterns_for_column(canonical)
        for pat in patterns:
            if SequenceMatcher(None, normed, pat).ratio() >= 0.85 and len(normed) > 3:
                return canonical
    return None


def _map_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    raw_cols = [str(c).strip().lower() for c in df.columns]
    new_names: list[str] = []
    used: set[str] = set()
    keep: list[bool] = []
    for idx, raw in enumerate(raw_cols):
        target = _best_column_match(raw)
        if target and target not in used:
            new_names.append(target)
            used.add(target)
            keep.append(True)
        elif target and target in used:
            new_names.append(raw)
            keep.append(False)
        else:
            new_names.append(str(df.columns[idx]))
            keep.append(True)
    df.columns = new_names
    df = df.loc[:, keep]
    return df


def _parse_numeric_scalar(val: Any) -> float:
    """Parse one cell: DE-style 1.234,56, US 1,234.56, plain floats, accounting (1.234)."""
    if val is None:
        return float("nan")
    if isinstance(val, bool):
        return float("nan")
    if isinstance(val, (int,)):
        return float(val)
    if isinstance(val, float):
        return float("nan") if pd.isna(val) else float(val)

    s = str(val).strip()
    for sym in ("\u20ac", "$", "\u00a3", "%"):
        s = s.replace(sym, "")
    s = s.replace("\xa0", " ").replace(" ", "").replace("'", "")
    if not s or s.lower() in ("nan", "none", "-", "—", "n/a"):
        return float("nan")

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()
    if s.endswith("-"):
        neg = not neg
        s = s[:-1].strip()
    if not s:
        return float("nan")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        if s.count(",") == 1 and re.search(r",\d{1,6}$", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]

    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return float("nan")


def _to_numeric_safe(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce").fillna(0.0)

    raw = series.astype(str)
    cleaned = (
        raw.str.replace("\u20ac", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace("\xa0", "", regex=False)
        .str.strip()
    )
    fast = pd.to_numeric(cleaned, errors="coerce")
    slow = series.map(_parse_numeric_scalar)
    merged = fast.where(fast.notna(), slow)
    return merged.fillna(0.0)


def _first_row_has_type_descriptors(df: pd.DataFrame) -> bool:
    """True if row 0 looks like Excel type-metadata (VarChar, Int, …)."""
    for val in df.iloc[0].values:
        vs = str(val).lower().strip()
        if vs in ("nan", "<na>", "nat", "none", ""):
            continue
        if any(kw in vs for kw in _TYPE_DESCRIPTOR_KEYWORDS):
            return True
    return False


def _drop_descriptor_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    if not _first_row_has_type_descriptors(df):
        return df
    rename_map: dict[str, str] = {}
    for col in df.columns:
        col_str = str(col)
        if "unnamed" not in col_str.lower() and not isinstance(col, (int, float)):
            continue
        desc_val = str(df.iloc[0][col]).strip()
        if desc_val and desc_val.lower() not in _TYPE_DESCRIPTOR_KEYWORDS and desc_val.lower() != "nan":
            rename_map[col] = desc_val
    if rename_map:
        df = df.rename(columns=rename_map)
    df = df.iloc[1:].reset_index(drop=True)
    return df


def _skip_title_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Skip rows that are titles/metadata (long text in first column, few useful columns)."""
    if df.empty or len(df) < 2:
        return df
    first_col = df.iloc[:, 0]
    for idx, val in enumerate(first_col):
        if idx > 5:
            break
        vs = str(val).strip() if pd.notna(val) else ""
        if not vs:
            continue
        if len(vs) > 50 and any(kw in vs.lower() for kw in ("rechnungswesen", "kanzlei", "kreditorenstammdaten", "kreditor")):
            return df.iloc[idx+1:].reset_index(drop=True)
    return df


def _safe_sort_filter_options(values: list[Any]) -> list[Any]:
    """Sort unique filter values; Excel columns may mix int/float/str."""
    try:
        return sorted(values)
    except TypeError:
        return sorted(values, key=lambda x: (type(x).__name__, str(x).casefold()))


def clean_dataframe(df: pd.DataFrame, drop_dup_rows: bool = True) -> pd.DataFrame:
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.loc[:, ~df.columns.duplicated()]
    if drop_dup_rows and len(df) <= 25_000:
        df = df.drop_duplicates().reset_index(drop=True)
    for col in df.columns:
        try:
            df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)
        except (AttributeError, TypeError, ValueError):
            pass
    return df


def _cell_str_lower(val: Any) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip().lower()
    if s in ("nan", "<na>", "nat", "none"):
        return ""
    return s


def _normalize_header_label(name: Any, col_index: int) -> str:
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return f"column_{col_index}"
    s = str(name).strip()
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    if not s or s.lower() == "nan":
        return f"column_{col_index}"
    slug = re.sub(r"[^\w\s\-\./&]", "_", s, flags=re.UNICODE)
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or f"column_{col_index}"


def _sanitize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    new_cols: list[str] = []
    seen: dict[str, int] = {}
    for i, c in enumerate(df.columns):
        base = _normalize_header_label(c, i)
        key = base.lower()
        cnt = seen.get(key, 0)
        seen[key] = cnt + 1
        col_name = base if cnt == 0 else f"{base}_{cnt}"
        new_cols.append(col_name)
    df.columns = new_cols
    return df


def _maybe_fill_merged_first_column(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or df.shape[1] < 1 or len(df) > 150_000:
        return df
    c0 = df.iloc[:, 0]
    try:
        null_ratio = float(c0.isna().mean())
    except Exception:
        return df
    if null_ratio < 0.12:
        return df
    try:
        sample = c0.dropna().head(20)
        if sample.empty:
            return df
        if pd.to_numeric(sample, errors="coerce").notna().mean() > 0.7:
            return df
    except Exception:
        pass
    out = df.copy()
    try:
        out.iloc[:, 0] = c0.ffill()
    except Exception:
        return df
    return out


def _detect_header_row(df_raw: pd.DataFrame, max_scan: int = 15) -> int:
    best_row, best_score = 0, 0
    n = min(max_scan, len(df_raw))
    ncols = df_raw.shape[1]
    for i in range(n):
        score = 0
        nonempty = 0
        for j in range(ncols):
            vs = _cell_str_lower(df_raw.iat[i, j])
            if not vs:
                continue
            nonempty += 1
            if _best_column_match(vs):
                score += 3
            elif len(vs) > 1 and not vs.replace(".", "").replace(",", "").replace("-", "").isdigit():
                score += 1
        if nonempty < 2:
            continue
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _excel_engines_for_read(filename: str) -> list[str]:
    """Prefer the right engine: .xls needs xlrd; .xlsx / .xlsm use openpyxl."""
    lower = (filename or "").lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        try:
            import xlrd  # noqa: F401

            return ["xlrd"]
        except ImportError:
            logger.warning("Install xlrd for .xls files: pip install xlrd")
            return ["openpyxl"]
    if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return ["openpyxl"]
    engines = ["openpyxl"]
    try:
        import xlrd  # noqa: F401

        engines.append("xlrd")
    except ImportError:
        pass
    return engines


def _read_excel_kw_variants(filename: str) -> list[dict[str, Any]]:
    variants: list[dict[str, Any]] = []
    for eng in _excel_engines_for_read(filename):
        variants.append({"engine": eng, "dtype": object})
        variants.append({"engine": eng})
    variants.append({"dtype": object})
    variants.append({})
    return variants


def _open_excel_file(raw_bytes: bytes, filename: str) -> pd.ExcelFile:
    bio = io.BytesIO(raw_bytes)
    last_err: Optional[Exception] = None
    for eng in _excel_engines_for_read(filename):
        try:
            bio.seek(0)
            return pd.ExcelFile(bio, engine=eng)
        except Exception as e:
            last_err = e
            logger.debug("ExcelFile engine=%s failed: %s", eng, e)
            bio = io.BytesIO(raw_bytes)
    try:
        bio.seek(0)
        return pd.ExcelFile(bio)
    except Exception as e:
        if last_err:
            raise last_err from e
        raise e


def _read_sheet_smart(raw_bytes: bytes, sheet_name: str, filename: str = "") -> pd.DataFrame:
    return _read_excel_sheet_robust(raw_bytes, sheet_name, filename)


def _read_excel_sheet_robust(raw_bytes: bytes, sheet_name: str, filename: str = "") -> pd.DataFrame:
    bio = io.BytesIO(raw_bytes)
    last_err: Optional[Exception] = None
    for extra in _read_excel_kw_variants(filename):
        try:
            bio.seek(0)
            df_peek = pd.read_excel(
                bio, sheet_name=sheet_name, header=None, nrows=25, **extra
            )
            header_row = _detect_header_row(df_peek)
            bio.seek(0)
            df = pd.read_excel(bio, sheet_name=sheet_name, header=header_row, **extra)
            if len(df) > _XLSX_ROW_WARN:
                logger.warning(
                    "Sheet %r is large (%s rows); processing may be slow.",
                    sheet_name,
                    len(df),
                )
            df = _sanitize_dataframe_columns(df)
            df = _maybe_fill_merged_first_column(df)
            df = _skip_title_rows(df)
            df = _drop_descriptor_rows(df)
            unnamed = sum(
                1
                for c in df.columns
                if str(c).lower().startswith(("unnamed", "column_"))
            )
            if (
                len(df.columns) > 0
                and unnamed > len(df.columns) * 0.65
                and header_row == 0
                and len(df) > 0
            ):
                for try_row in range(1, min(8, len(df))):
                    hits = 0
                    for j in range(min(df.shape[1], 40)):
                        vs = _cell_str_lower(df.iat[try_row - 1, j])
                        if vs and _best_column_match(vs):
                            hits += 1
                    if hits >= 2:
                        bio.seek(0)
                        df = pd.read_excel(bio, sheet_name=sheet_name, header=try_row, **extra)
                        df = _sanitize_dataframe_columns(df)
                        df = _maybe_fill_merged_first_column(df)
                        df = _skip_title_rows(df)
                        df = _drop_descriptor_rows(df)
                        break
            return df
        except Exception as e:
            last_err = e
            continue
    logger.warning("Could not read sheet %r: %s", sheet_name, last_err)
    return pd.DataFrame()


def _iter_excel_sheet_names(raw_bytes: bytes, xls: pd.ExcelFile) -> list[str]:
    names: list[str] = []
    try:
        from openpyxl import load_workbook

        wb = load_workbook(
            io.BytesIO(raw_bytes),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            names = list(wb.sheetnames)
        finally:
            wb.close()
    except Exception as e:
        logger.debug("openpyxl sheet enumeration: %s", e)
    if not names:
        return list(xls.sheet_names)
    for s in xls.sheet_names:
        if s not in names:
            names.append(s)
    return names


def _read_excel_no_header(raw_bytes: bytes, sheet_name: str, filename: str = "") -> pd.DataFrame:
    """Raw grid (integer columns); used where parsers select columns by index."""
    bio = io.BytesIO(raw_bytes)
    for extra in _read_excel_kw_variants(filename):
        try:
            bio.seek(0)
            return pd.read_excel(bio, sheet_name=sheet_name, header=None, **extra)
        except Exception:
            continue
    return pd.DataFrame()


def _classify_sheet(sheet_name: str, df: pd.DataFrame) -> str:
    name = sheet_name.strip().lower()
    if "material" in name:
        return "transactions"
    if "kreditor" in name:
        return "suppliers"
    if "pivot" in name or "tabelle" in name:
        return "pivot"
    if "verbindlich" in name or "liabilit" in name:
        return "liabilities"
    if "zusammenfassung" in name or "summary" in name:
        return "summary"
    cols_lower = {str(c).strip().lower() for c in df.columns}
    tx_signals = {"debit", "credit", "soll", "haben", "umsatz soll", "umsatz haben",
                  "period", "periode", "bundlecode", "financialpostingcode",
                  "betrag", "amount", "netto", "brutto"}
    if len(cols_lower & tx_signals) >= 2:
        return "transactions"
    supplier_signals = {"beschriftung", "kurzbezeichnung", "plz", "ort",
                        "supplier name", "kreditorenstammdaten", "telefon",
                        "email", "adresse", "address"}
    first_col_str = str(list(df.columns)[0]).lower() if len(df.columns) > 0 else ""
    if len(cols_lower & supplier_signals) >= 1 or "kreditorenstammdaten" in first_col_str:
        return "suppliers"
    if df.shape[1] == 2 and df.shape[0] < 500:
        return "pivot"
    if len(_detect_numeric_columns(df)) >= 1 and len(df) >= 3:
        return "transactions"
    return "unknown"


def _detect_numeric_columns(df: pd.DataFrame) -> list[str]:
    result = []
    for col in df.columns:
        if df[col].dtype in ("float64", "int64", "float32", "int32"):
            if df[col].notna().sum() > len(df) * 0.1:
                result.append(str(col))
            continue
        try:
            parsed = df[col].map(_parse_numeric_scalar)
            rate = parsed.notna().sum() / max(len(df), 1)
            if rate > 0.3 and parsed.fillna(0.0).abs().sum() > 0:
                result.append(str(col))
        except Exception:
            pass
    return result


def _detect_date_columns(df: pd.DataFrame) -> list[str]:
    result = []
    for col in df.columns:
        if df[col].dtype == "datetime64[ns]":
            result.append(str(col))
            continue
        try:
            sample = df[col].dropna().head(50)
            if sample.empty:
                continue
            parsed = pd.to_datetime(sample.astype(str), dayfirst=True, errors="coerce")
            if parsed.notna().sum() > len(sample) * 0.5:
                result.append(str(col))
        except Exception:
            pass
    return result


def _detect_category_columns(df: pd.DataFrame) -> list[str]:
    result = []
    for col in df.columns:
        if col in (C_DATE, C_PERIOD, C_DEBIT, C_CREDIT, C_AMOUNT):
            continue
        try:
            nunique = df[col].nunique()
            if 2 <= nunique <= min(200, len(df) * 0.5) and df[col].dtype == "object":
                result.append(str(col))
        except Exception:
            pass
    return result


def _parse_period_to_date(series: pd.Series, year: int = 2025) -> pd.Series:
    dates = []
    for val in series:
        val_str = str(val).strip().lower()
        if val_str in _MONTH_NAMES_DE:
            dates.append(pd.Timestamp(year=year, month=_MONTH_NAMES_DE[val_str], day=1))
            continue
        try:
            num = int(float(val_str))
            if 1 <= num <= 12:
                dates.append(pd.Timestamp(year=year, month=num, day=1))
            elif 202000 <= num <= 209912:
                y, m = divmod(num, 100)
                dates.append(pd.Timestamp(year=y, month=m, day=1) if 1 <= m <= 12 else pd.NaT)
            else:
                dates.append(pd.NaT)
        except (ValueError, TypeError):
            try:
                dates.append(pd.to_datetime(val_str, dayfirst=True))
            except Exception:
                dates.append(pd.NaT)
    return pd.Series(dates, index=series.index)


def _extract_year_from_filename(filename: str) -> int:
    match = re.search(r"(20\d{2})", filename)
    return int(match.group(1)) if match else 2025


def _parse_transactions(df: pd.DataFrame, filename: str = "") -> pd.DataFrame:
    df = _drop_descriptor_rows(df.copy())
    df = _map_columns(df)
    df = clean_dataframe(df, drop_dup_rows=False)
    if df.empty:
        return pd.DataFrame()
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].astype(str)
    file_year = _extract_year_from_filename(filename)

    if C_PERIOD in df.columns and C_DATE not in df.columns:
        df[C_DATE] = _parse_period_to_date(df[C_PERIOD], year=file_year)
    elif C_DATE in df.columns:
        if df[C_DATE].dtype == "object" or df[C_DATE].dtype == "O":
            df[C_DATE] = pd.to_datetime(df[C_DATE], dayfirst=True, errors="coerce")
    if C_DATE not in df.columns:
        dc = _detect_date_columns(df)
        if dc:
            df[C_DATE] = pd.to_datetime(df[dc[0]], dayfirst=True, errors="coerce")
        else:
            df[C_DATE] = pd.NaT

    for col in (C_DEBIT, C_CREDIT):
        if col in df.columns:
            df[col] = _to_numeric_safe(df[col])
        else:
            df[col] = 0.0

    debit_looks_wrong = (
        df[C_DEBIT].abs().max() <= 12 and df[C_DEBIT].nunique() <= 13 and len(df) > 20
    )

    if C_AMOUNT in df.columns:
        summe = _to_numeric_safe(df[C_AMOUNT])
        if debit_looks_wrong or (df[C_DEBIT].sum() == 0 and df[C_CREDIT].sum() == 0):
            df[C_DEBIT] = summe.clip(lower=0)
            df[C_CREDIT] = summe.clip(upper=0).abs()
    else:
        if df[C_DEBIT].sum() == 0 and df[C_CREDIT].sum() == 0:
            nc = _detect_numeric_columns(df)
            non_canon = [c for c in nc if c not in {C_DEBIT, C_CREDIT, C_AMOUNT, C_SUPPLIER_ID} and c in df.columns]
            if non_canon:
                av = _to_numeric_safe(df[non_canon[0]])
                df[C_DEBIT] = av.clip(lower=0)
                df[C_CREDIT] = av.clip(upper=0).abs()

    df[C_AMOUNT] = df[C_DEBIT] - df[C_CREDIT]

    if C_SUPPLIER_ID in df.columns:
        df[C_SUPPLIER_ID] = pd.to_numeric(
            df[C_SUPPLIER_ID].astype(str).str.replace("C_", "", regex=False), errors="coerce"
        )

    if C_SUPPLIER_NAME not in df.columns:
        if C_DESCRIPTION in df.columns:
            df[C_SUPPLIER_NAME] = df[C_DESCRIPTION]
        else:
            cc = _detect_category_columns(df)
            tc = [c for c in cc if c not in {C_DATE, C_PERIOD, C_DEBIT, C_CREDIT, C_AMOUNT, C_SUPPLIER_ID, C_GL_ACCOUNT, C_COST_CENTER}]
            df[C_SUPPLIER_NAME] = df[tc[0]] if tc else "Unknown"
    df[C_SUPPLIER_NAME] = df[C_SUPPLIER_NAME].fillna("Unknown").replace("nan", "Unknown")

    for col in (C_DESCRIPTION, C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY):
        if col not in df.columns:
            df[col] = ""

    keep = [C_DATE, C_SUPPLIER_ID, C_SUPPLIER_NAME, C_DEBIT, C_CREDIT, C_AMOUNT,
            C_DESCRIPTION, C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY]
    extras = [c for c in df.columns if c not in keep and c != C_PERIOD and not str(c).startswith("Unnamed")]
    result = df[[c for c in keep if c in df.columns] + extras].reset_index(drop=True)
    result = result[(result[C_DEBIT] != 0) | (result[C_CREDIT] != 0) | (result[C_AMOUNT] != 0)]
    return result.reset_index(drop=True)


def _parse_suppliers(df_raw: pd.DataFrame, raw_bytes: bytes, sheet_name: str, filename: str = "") -> pd.DataFrame:
    first_col = str(list(df_raw.columns)[0]).lower()
    if "kreditorenstammdaten" in first_col or "rechnungswesen" in first_col:
        header_row = df_raw.iloc[0]
        cols = [str(v).strip() if pd.notna(v) else f"col_{i}" for i, v in enumerate(header_row)]
        df = df_raw.iloc[1:].copy()
        df.columns = cols
        df = clean_dataframe(df)
        id_col, name_col = None, None
        cols_lower = {c.lower(): c for c in df.columns}
        for key in ["konto", "nr", "nummer", "kreditoren nr"]:
            if key in cols_lower:
                id_col = cols_lower[key]
                break
        for key in ["beschriftung", "name", "kurzbezeichnung"]:
            if key in cols_lower:
                name_col = cols_lower[key]
                break
        if id_col is None:
            id_col = df.columns[0]
        if name_col is None:
            name_col = df.columns[2] if len(df.columns) > 2 else df.columns[min(1, len(df.columns) - 1)]
        result = pd.DataFrame({
            C_SUPPLIER_ID: pd.to_numeric(df[id_col].astype(str).str.replace(" ", ""), errors="coerce"),
            C_SUPPLIER_NAME: df[name_col].astype(str).str.strip(),
        })
        result = result.dropna(subset=[C_SUPPLIER_ID])
        result[C_SUPPLIER_ID] = result[C_SUPPLIER_ID].astype(int)
        result[C_SUPPLIER_NAME] = result[C_SUPPLIER_NAME].replace("nan", "Unknown")
        return result.drop_duplicates(subset=[C_SUPPLIER_ID]).reset_index(drop=True)

    all_unnamed = all("unnamed" in str(c).lower() or isinstance(c, (int, float)) for c in df_raw.columns)
    str_cols = [str(c).strip().lower() for c in df_raw.columns]
    hdr_kw = {"nr", "nummer", "name", "bezeichnung", "kreditoren", "supplier", "lieferant", "id", "konto", "beschriftung"}
    has_real = any(set(c.replace("-", " ").replace("_", " ").split()) & hdr_kw for c in str_cols if not c.startswith("unnamed"))

    if all_unnamed or not has_real:
        df_nh = _read_excel_no_header(raw_bytes, sheet_name, filename)
        if df_nh.empty:
            df_nh = df_raw.copy()
            df_nh.columns = range(len(df_nh.columns))
        id_col_idx, name_col_idx, company_col_idx = None, None, None
        for idx in range(min(df_nh.shape[1], 10)):
            sample = df_nh[idx].dropna()
            if sample.empty:
                continue
            is_numeric = pd.to_numeric(sample, errors="coerce").notna().mean() > 0.8
            has_dash = sample.astype(str).str.contains(" - ").mean() > 0.3
            if has_dash and company_col_idx is None:
                company_col_idx = idx
                continue
            if is_numeric and id_col_idx is None:
                nums = pd.to_numeric(sample, errors="coerce").dropna()
                if len(nums) > 0 and nums.mean() > 1000:
                    id_col_idx = idx
                    continue
            if not is_numeric and name_col_idx is None and sample.dtype == object:
                name_col_idx = idx
                continue
        if id_col_idx is None:
            id_col_idx = 3 if df_nh.shape[1] > 3 else 0
        if name_col_idx is None:
            name_col_idx = 4 if df_nh.shape[1] > 4 else 1
        result = pd.DataFrame({
            C_SUPPLIER_ID: pd.to_numeric(df_nh[id_col_idx], errors="coerce"),
            C_SUPPLIER_NAME: df_nh[name_col_idx].astype(str).str.strip(),
        })
        if company_col_idx is not None:
            result[C_COMPANY] = df_nh[company_col_idx].apply(
                lambda v: str(v).split("-", 1)[1].strip() if isinstance(v, str) and "-" in v else str(v))
        result = result.dropna(subset=[C_SUPPLIER_ID])
        result[C_SUPPLIER_ID] = result[C_SUPPLIER_ID].astype(int)
        result[C_SUPPLIER_NAME] = result[C_SUPPLIER_NAME].replace("nan", "Unknown")
        return result.drop_duplicates(subset=[C_SUPPLIER_ID]).reset_index(drop=True)

    df = _map_columns(df_raw.copy())
    df = clean_dataframe(df)
    if C_SUPPLIER_ID not in df.columns:
        df = df.rename(columns={df.columns[0]: C_SUPPLIER_ID})
    if C_SUPPLIER_NAME not in df.columns:
        for c in df.columns:
            if c != C_SUPPLIER_ID:
                df = df.rename(columns={c: C_SUPPLIER_NAME})
                break
    if C_SUPPLIER_NAME not in df.columns:
        df[C_SUPPLIER_NAME] = "Unknown"
    df[C_SUPPLIER_ID] = pd.to_numeric(df[C_SUPPLIER_ID], errors="coerce")
    df = df.dropna(subset=[C_SUPPLIER_ID])
    df[C_SUPPLIER_ID] = df[C_SUPPLIER_ID].astype(int)
    keep = [c for c in [C_SUPPLIER_ID, C_SUPPLIER_NAME, C_COMPANY] if c in df.columns]
    return df[keep].drop_duplicates(subset=[C_SUPPLIER_ID]).reset_index(drop=True)


def _parse_pivot(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all").reset_index(drop=True)
    for i in range(min(5, len(df))):
        row_str = df.iloc[i].astype(str).str.lower()
        if row_str.str.contains("zeilenbeschriftung|summe von|supplier|manufacturer", regex=True).any():
            df = df.iloc[i + 1:].reset_index(drop=True)
            break
    if df.shape[1] < 2:
        return pd.DataFrame(columns=[C_SUPPLIER_NAME, C_AMOUNT])
    result = pd.DataFrame({
        C_SUPPLIER_NAME: df[df.columns[0]].astype(str).str.strip(),
        C_AMOUNT: _to_numeric_safe(df[df.columns[1]]),
    })
    result = result[result[C_SUPPLIER_NAME].notna() & (result[C_SUPPLIER_NAME] != "nan")]
    return result[result[C_AMOUNT].abs() > 0].reset_index(drop=True)


def _parse_generic(df: pd.DataFrame, filename: str = "") -> pd.DataFrame:
    df = _map_columns(df.copy())
    df = _drop_descriptor_rows(df)
    df = clean_dataframe(df, drop_dup_rows=False)
    if df.empty or len(df.columns) < 2:
        return pd.DataFrame()
    nc = _detect_numeric_columns(df)
    if not nc:
        return pd.DataFrame()
    for col in nc:
        if col in df.columns:
            df[col] = _to_numeric_safe(df[col])
    if C_AMOUNT not in df.columns:
        best = next((c for c in nc if c not in {C_SUPPLIER_ID, C_DEBIT, C_CREDIT}), None)
        df[C_AMOUNT] = _to_numeric_safe(df[best]) if best else 0.0
    if C_DEBIT not in df.columns:
        df[C_DEBIT] = df[C_AMOUNT].clip(lower=0) if C_AMOUNT in df.columns else 0.0
    if C_CREDIT not in df.columns:
        df[C_CREDIT] = df[C_AMOUNT].clip(upper=0).abs() if C_AMOUNT in df.columns else 0.0
    if C_DATE not in df.columns:
        dc = _detect_date_columns(df)
        if dc:
            df[C_DATE] = pd.to_datetime(df[dc[0]], dayfirst=True, errors="coerce")
        elif C_PERIOD in df.columns:
            df[C_DATE] = _parse_period_to_date(df[C_PERIOD], year=_extract_year_from_filename(filename))
        else:
            df[C_DATE] = pd.NaT
    if C_SUPPLIER_NAME not in df.columns:
        cc = _detect_category_columns(df)
        tc = [c for c in cc if c not in set(nc)]
        if tc:
            df[C_SUPPLIER_NAME] = df[tc[0]]
        elif C_DESCRIPTION in df.columns:
            df[C_SUPPLIER_NAME] = df[C_DESCRIPTION]
        else:
            df[C_SUPPLIER_NAME] = "Unknown"
    df[C_SUPPLIER_NAME] = df[C_SUPPLIER_NAME].fillna("Unknown").replace("nan", "Unknown")
    for col in (C_DESCRIPTION, C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY):
        if col not in df.columns:
            df[col] = ""
    keep = [C_DATE, C_SUPPLIER_ID, C_SUPPLIER_NAME, C_DEBIT, C_CREDIT, C_AMOUNT,
            C_DESCRIPTION, C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY]
    return df[[c for c in keep if c in df.columns]].reset_index(drop=True)


_INVOICE_SUFFIXES = (
    "_rechnungseingang", "_rechnungseingan", "_rechnungseing",
    "_e-rechnung", "_rechnung", "_rechnungseing",
    "rechnungseingang", "rechnungseing",
)

_LEGAL_SUFFIXES = (
    " gmbh & co. kg", " gmbh & co.kg", " gmbh & co kg",
    " gmbh & co", " gmbh", " ag", " kg", " ohg", " e.k.",
    " e.k", " mbh", " ug", " se", " eg", " co.", " inc.", " ltd.",
    " corp.", " s.a.", " s.r.l.",
)

# Hyphen, en dash, em dash — after "621114 - Supplier GmbH" style labels
_KONTO_PREFIX_RE = re.compile(r"^\d{5,12}\s*[\u002d\u2013\u2014]\s*")


def _strip_leading_konto_prefix(n: str) -> str:
    """Drop leading Konto/Kreditor number + dash so '608238 - Firma' and '621114 - Firma' share one key."""
    return _KONTO_PREFIX_RE.sub("", n, count=1)


def _german_ascii_fold(s: str) -> str:
    """Lowercase, expand umlauts, strip accents (Müller/Mueller/Muller align better)."""
    s = s.strip().lower()
    repl = {
        "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
        "à": "a", "á": "a", "â": "a", "ã": "a",
        "è": "e", "é": "e", "ê": "e", "ë": "e",
        "ì": "i", "í": "i", "î": "i", "ï": "i",
        "ò": "o", "ó": "o", "ô": "o", "õ": "o",
        "ù": "u", "ú": "u", "û": "u",
        "ý": "y", "ÿ": "y", "ñ": "n", "ç": "c",
    }
    s = "".join(repl.get(c, c) for c in s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _skonto_konto_from_label(n: str) -> Optional[re.Match]:
    """Match 'Skontoabzug 80076 / …' style lines; same Konto → one merge key."""
    return re.match(r"^(skontoabzug)\s+(\d{4,8})\b", n.strip())


def _canonical_supplier_label_for_key(match_key: str, members: list[str]) -> str:
    """Prefer short Skontoabzug <Konto> labels; else longest original name."""
    if match_key and re.fullmatch(r"skontoabzug \d{4,8}", match_key):
        num = match_key.split()[-1]
        return f"Skontoabzug {num}"
    if not members:
        return "Unknown"
    return max(members, key=lambda x: (len(str(x)), str(x)))


def _supplier_match_key(name: str) -> str:
    """Stable key for grouping: strips booking refs, Skontoabzug line IDs, invoice tails."""
    if not name or str(name).strip().lower() in ("", "nan", "unknown", "none"):
        return ""
    n = _german_ascii_fold(str(name))
    n = re.sub(r"\s*,\s*", " ", n)
    n = re.sub(r"\s+eg\s+", " ", n)
    n = _strip_leading_konto_prefix(n)

    sk0 = _skonto_konto_from_label(n)
    if sk0:
        return f"{sk0.group(1)} {sk0.group(2)}"

    for suf in _LEGAL_SUFFIXES:
        if n.endswith(suf):
            n = n[: -len(suf)].strip()

    n = re.sub(r"\s*#\s*\d+.*$", "", n)
    n = re.sub(r"\s+proj\.?\s*_.*$", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+wu\s*$", "", n, flags=re.IGNORECASE)

    for suf in _INVOICE_SUFFIXES:
        if suf in n:
            n = n[: n.index(suf)].strip()
            break

    n = re.sub(r"\s+/\s*hn-v\s+.*$", "", n, flags=re.IGNORECASE)
    n = re.sub(r"\s+/\s*\d[\d\s]*$", "", n)
    n = re.sub(r"\s+/\s*\d+$", "", n)

    sk = _skonto_konto_from_label(n)
    if sk:
        return f"{sk.group(1)} {sk.group(2)}"

    hm = re.search(r"hn-?\s*v\s*(\d+)", n.replace(" ", ""))
    if hm:
        return f"hnv{hm.group(1)}"

    n = re.sub(r"[/\\]+", " ", n)
    n = re.sub(r"[_]+", " ", n)
    n = re.sub(r"[^\w\s\-]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    n = re.sub(r"\s+\d{6,}\s*$", "", n)
    return n


def _normalize_supplier_name(name: str) -> str:
    """Strip legal suffixes; used as secondary signal with match key."""
    n = _german_ascii_fold(str(name))
    for suffix in _LEGAL_SUFFIXES:
        if n.endswith(suffix):
            n = n[: -len(suffix)].strip()
    n = re.sub(r"[,./&\-]+$", "", n).strip()
    n = re.sub(r"\s+", " ", n)
    return n


def _name_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    ra = SequenceMatcher(None, a, b).ratio()
    rb = 0.0
    la, lb = len(a), len(b)
    shorter, longer = (a, b) if la <= lb else (b, a)
    if len(shorter) >= 4 and len(longer) >= 4:
        if longer.startswith(shorter) or longer.endswith(" " + shorter):
            rb = max(rb, 0.92)
    if len(shorter) >= 6 and len(longer) >= len(shorter) + 2:
        if re.search(r"(^|\s)" + re.escape(shorter) + r"(\s|$)", longer):
            rb = max(rb, 0.91)
    tokens_a = set(a.split())
    tokens_b = set(b.split())
    if tokens_a and tokens_b:
        inter = len(tokens_a & tokens_b)
        union = len(tokens_a | tokens_b)
        if union > 0:
            rb = max(rb, inter / union)
        if len(shorter) >= 5:
            if shorter in tokens_a and shorter in tokens_b:
                rb = max(rb, 0.93)
            elif shorter in tokens_a or shorter in tokens_b:
                longer_tokens = tokens_b if shorter in tokens_a else tokens_a
                for t in longer_tokens:
                    if len(t) >= len(shorter) and t != shorter and shorter in t and len(shorter) >= 6:
                        rb = max(rb, 0.90)
                        break
    return max(ra, rb)


def _collapse_names_by_match_key(names: list[str]) -> dict[str, str]:
    """Map each raw supplier label to one canonical name per _supplier_match_key bucket (longest wins)."""
    raw = sorted(
        {
            str(n).strip()
            for n in names
            if str(n).strip() and str(n).lower() not in ("nan", "unknown", "none")
        }
    )
    if not raw:
        return {}
    buckets: dict[str, list[str]] = {}
    for name in raw:
        k = _supplier_match_key(name)
        if not k:
            k = _normalize_supplier_name(name) or name.lower()
        buckets.setdefault(k, []).append(name)

    mapping: dict[str, str] = {}
    for k, members in buckets.items():
        canon = _canonical_supplier_label_for_key(k, members)
        for m in members:
            mapping[m] = canon
    return mapping


def _series_apply_name_map(ser: pd.Series, mapping: dict[str, str]) -> pd.Series:
    """Apply supplier rename dict; keep original where key missing (vectorized)."""
    if not mapping:
        return ser
    s = ser.fillna("Unknown").astype(str).str.strip()
    hit = s.map(mapping)
    return hit.combine_first(s)


def _fuzzy_group_names(names: list[str], threshold: float = FUZZY_SUPPLIER_MERGE_THRESHOLD) -> dict[str, str]:
    """Map similar supplier names to one canonical display name (longest original).

    Pairs of internal match-keys must reach ``threshold`` (default 88%) similarity to merge;
    all rows under merged names roll up into one line when amounts are aggregated.
    """
    raw = sorted({str(n).strip() for n in names if str(n).strip() and str(n).lower() not in ("nan", "unknown")})
    if len(raw) <= 1:
        return {n: n for n in raw}

    key_for: dict[str, str] = {}
    buckets: dict[str, list[str]] = {}
    for name in raw:
        k = _supplier_match_key(name)
        if not k:
            k = _normalize_supplier_name(name) or name.lower()
        key_for[name] = k
        buckets.setdefault(k, []).append(name)

    def canonical_for_bucket_key(k: str, members: list[str]) -> str:
        return _canonical_supplier_label_for_key(k, members)

    key_canon: dict[str, str] = {
        k: canonical_for_bucket_key(k, v) for k, v in buckets.items()
    }
    keys = list(key_canon.keys())

    _FUZZY_KEY_CAP = 700
    if len(keys) > _FUZZY_KEY_CAP:
        logger.info(
            "Skipping cross-bucket fuzzy merge (%s match-keys); using per-bucket names only.",
            len(keys),
        )
        mapping: dict[str, str] = {}
        for k, members in buckets.items():
            canon = canonical_for_bucket_key(k, members)
            for m in members:
                mapping[m] = canon
        return mapping

    parent = {k: k for k in keys}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x: str, y: str) -> None:
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py

    def _structured_keys_distinct(a: str, b: str) -> bool:
        """Do not fuzzy-merge different booking accounts (Skontoabzug / HN-V IDs)."""
        if re.match(r"^skontoabzug \d+$", a) and re.match(r"^skontoabzug \d+$", b):
            return a != b
        if re.match(r"^hnv\d+$", a) and re.match(r"^hnv\d+$", b):
            return a != b
        return False

    short_thr = max(0.75, threshold - 0.10)
    for i, ka in enumerate(keys):
        for kb in keys[i + 1 :]:
            if ka == kb:
                continue
            if _structured_keys_distinct(ka, kb):
                continue
            ca, cb = key_canon[ka], key_canon[kb]
            sim_key = _name_similarity(ka, kb)
            sim_norm = _name_similarity(
                _normalize_supplier_name(ca), _normalize_supplier_name(cb)
            )
            min_len = min(len(ka), len(kb))
            thr = short_thr if min_len <= 12 else threshold
            if sim_key >= thr or sim_norm >= thr:
                union(ka, kb)
                continue
            if min_len >= 5 and (ka.startswith(kb + " ") or kb.startswith(ka + " ")):
                union(ka, kb)

    root_members: dict[str, list[str]] = {}
    for k in keys:
        r = find(k)
        root_members.setdefault(r, []).extend(buckets[k])

    mapping: dict[str, str] = {}
    for _root, members in root_members.items():
        uniq = list(dict.fromkeys(members))
        roots = {_supplier_match_key(m) or _normalize_supplier_name(m) or m.lower() for m in uniq}
        roots.discard("")
        rk = next(iter(roots)) if len(roots) == 1 else ""
        canon = _canonical_supplier_label_for_key(rk, uniq) if rk else max(uniq, key=lambda x: (len(x), x))
        for m in members:
            mapping[m] = canon

    return mapping


def _merge_transactions_suppliers(transactions: pd.DataFrame, suppliers: pd.DataFrame) -> pd.DataFrame:
    if suppliers.empty or C_SUPPLIER_ID not in transactions.columns or C_SUPPLIER_ID not in suppliers.columns:
        merged = transactions.copy()
    else:
        sup_merge = suppliers[[c for c in [C_SUPPLIER_ID, C_SUPPLIER_NAME, C_COMPANY] if c in suppliers.columns]].drop_duplicates(subset=[C_SUPPLIER_ID])
        if C_SUPPLIER_NAME in sup_merge.columns:
            sup_merge = sup_merge.rename(columns={C_SUPPLIER_NAME: "_SupName"})
        merged = transactions.merge(sup_merge, on=C_SUPPLIER_ID, how="left")
        if "_SupName" in merged.columns:
            # Check which rows have missing/invalid supplier names
            current_str = merged[C_SUPPLIER_NAME].fillna("").astype(str).str.strip()

            # Detect numeric IDs: pure numbers with optional dots/dashes
            is_numeric_id = current_str.apply(lambda x: bool(x) and x.replace(".", "").replace("-", "").isdigit())

            # Detect missing/unknown values
            is_missing = current_str.str.lower().isin(["unknown", "nan", "", "none", "<na>"])

            # Replace where appropriate
            mask = (is_missing | is_numeric_id) & merged["_SupName"].notna()
            merged.loc[mask, C_SUPPLIER_NAME] = merged.loc[mask, "_SupName"]

            merged.drop(columns=["_SupName"], inplace=True)
    if C_COMPANY not in merged.columns:
        merged[C_COMPANY] = "Unknown"
    merged[C_COMPANY] = merged[C_COMPANY].fillna("Unknown")
    merged[C_SUPPLIER_NAME] = merged[C_SUPPLIER_NAME].fillna("Unknown")
    return merged


def _extract_company_from_filename(filename: str) -> str:
    name = filename.replace(".xlsx", "").replace(".xls", "").replace(".csv", "")
    name = re.sub(r"^Kopie von\s*", "", name, flags=re.IGNORECASE)
    name = re.sub(r"^Copy of\s*", "", name, flags=re.IGNORECASE)
    name = re.sub(r"^\d{6}_", "", name)
    name = re.sub(r"_Materialkosten.*$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"_Material.*$", "", name, flags=re.IGNORECASE)
    return name.replace("_", " ").strip() or filename


def process_file(uploaded) -> dict:
    _load_reference_schema_aliases()
    raw_bytes = uploaded.getvalue()
    filename = uploaded.name
    company_label = _extract_company_from_filename(filename)
    all_transactions: list[pd.DataFrame] = []
    all_suppliers: list[pd.DataFrame] = []
    all_pivots: list[pd.DataFrame] = []
    sheet_info: list[dict] = []
    is_csv = filename.lower().endswith(".csv")

    if is_csv:
        df_raw = pd.DataFrame()
        for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                df_raw = pd.read_csv(
                    io.BytesIO(raw_bytes),
                    encoding=enc,
                    sep=None,
                    engine="python",
                    on_bad_lines="skip",
                )
                break
            except TypeError:
                try:
                    df_raw = pd.read_csv(
                        io.BytesIO(raw_bytes), encoding=enc, sep=None, engine="python"
                    )
                    break
                except Exception:
                    continue
            except Exception:
                continue
        if not df_raw.empty:
            try:
                df_raw = _sanitize_dataframe_columns(df_raw)
            except Exception as e:
                logger.warning("CSV column sanitize skipped: %s", e)
            classification = _classify_sheet("data", df_raw)
            sheet_info.append({"sheet": "CSV Data", "type": classification, "rows": len(df_raw), "cols": len(df_raw.columns)})
            try:
                parsed = _parse_transactions(df_raw, filename)
                if not parsed.empty:
                    parsed[C_SOURCE] = company_label
                    all_transactions.append(parsed)
            except Exception as e:
                logger.warning("CSV parse failed for %s: %s", filename, e)
                sheet_info[-1]["note"] = str(e)
    else:
        try:
            xls = _open_excel_file(raw_bytes, filename)
        except Exception as e:
            logger.warning("ExcelFile open failed for %s: %s", filename, e)
            hint = ""
            if "encrypted" in str(e).lower() or "password" in str(e).lower():
                hint = " (Workbook may be password-protected — remove protection in Excel and save.)"
            elif filename.lower().endswith(".xls") and "xlrd" not in str(e).lower():
                hint = " For .xls install: pip install xlrd"
            return {"filename": filename, "company": company_label, "transactions": pd.DataFrame(),
                    "suppliers": pd.DataFrame(), "pivots": pd.DataFrame(),
                    "sheet_info": [{"sheet": "ERROR", "type": "unreadable", "rows": 0, "cols": 0,
                                    "note": str(e) + hint}]}
        for sheet_name in _iter_excel_sheet_names(raw_bytes, xls):
            try:
                df_raw = _read_excel_sheet_robust(raw_bytes, sheet_name, filename)
            except Exception as e:
                logger.warning("Sheet %r read failed (%s): %s", sheet_name, filename, e)
                sheet_info.append({
                    "sheet": sheet_name, "type": "unreadable", "rows": 0, "cols": 0, "note": str(e),
                })
                continue
            if df_raw.empty or df_raw.shape[1] == 0:
                sheet_info.append({
                    "sheet": sheet_name, "type": "empty", "rows": 0, "cols": 0,
                })
                continue
            try:
                classification = _classify_sheet(sheet_name, df_raw)
            except Exception as e:
                logger.warning("classify_sheet %r: %s", sheet_name, e)
                classification = "unknown"
            sheet_info.append({"sheet": sheet_name, "type": classification, "rows": len(df_raw), "cols": len(df_raw.columns)})
            try:
                if classification == "transactions":
                    parsed = _parse_transactions(df_raw, filename)
                    if not parsed.empty:
                        parsed[C_SOURCE] = company_label
                        parsed[C_SHEET] = sheet_name
                        all_transactions.append(parsed)
                elif classification == "suppliers":
                    parsed = _parse_suppliers(df_raw, raw_bytes, sheet_name, filename)
                    all_suppliers.append(parsed)
                elif classification == "pivot":
                    parsed = _parse_pivot(df_raw)
                    if not parsed.empty:
                        parsed[C_SOURCE] = company_label
                        all_pivots.append(parsed)
                elif classification in ("liabilities", "summary"):
                    parsed = _parse_transactions(df_raw, filename)
                    if not parsed.empty:
                        parsed[C_SOURCE] = company_label
                        parsed[C_SHEET] = sheet_name
                        all_transactions.append(parsed)
                else:
                    parsed = _parse_generic(df_raw, filename)
                    if not parsed.empty and len(parsed) >= 2:
                        parsed[C_SOURCE] = company_label
                        parsed[C_SHEET] = sheet_name
                        all_transactions.append(parsed)
            except Exception as e:
                logger.warning("Parse failed sheet %r in %s: %s", sheet_name, filename, e)
                sheet_info[-1]["note"] = str(e)

    transactions = pd.concat(all_transactions, ignore_index=True) if all_transactions else pd.DataFrame()
    suppliers = pd.concat(all_suppliers, ignore_index=True) if all_suppliers else pd.DataFrame()
    pivots = pd.concat(all_pivots, ignore_index=True) if all_pivots else pd.DataFrame()
    if not transactions.empty and not suppliers.empty:
        transactions = _merge_transactions_suppliers(transactions, suppliers)
    return {"filename": filename, "company": company_label, "transactions": transactions,
            "suppliers": suppliers, "pivots": pivots, "sheet_info": sheet_info}


def _agg_join_unique_strings(series: pd.Series) -> str:
    """Join unique non-empty string values in a group (Arrow/pd.NA safe)."""
    seen: set[str] = set()
    for val in series:
        if pd.isna(val):
            continue
        s = str(val).strip()
        if not s or s.lower() == "nan":
            continue
        seen.add(s)
    return ", ".join(sorted(seen))


def _agg_join_months(series: pd.Series) -> str:
    seen: set[str] = set()
    for val in series:
        if pd.isna(val):
            continue
        s = str(val).strip()
        if not s or s in ("N/A", "nan"):
            continue
        for part in s.split(","):
            p = part.strip()
            if p and p not in ("N/A", "nan"):
                seen.add(p)
    return ", ".join(sorted(seen))


def _agg_join_unique_ids(series: pd.Series) -> str:
    seen: set[str] = set()
    for val in series:
        if pd.isna(val):
            continue
        try:
            iv = int(float(val))
            seen.add(str(iv))
        except (ValueError, TypeError):
            s = str(val).strip()
            if s and s.lower() != "nan":
                seen.add(s)
    return ", ".join(sorted(seen, key=lambda x: (len(x), x)))


def _normalize_supplier_group_key(name: str) -> str:
    """Stable key: trim, NFKC, collapse spaces, case-insensitive (casefold). Empty if unknown."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = unicodedata.normalize("NFKC", str(name)).strip()
    if not s or s.lower() in ("nan", "none", "unknown", ""):
        return ""
    s = _strip_leading_konto_prefix(s)
    s = re.sub(r"\s*,\s*", " ", s)
    s = re.sub(r"\s+", " ", s).strip().casefold()
    return s


def _consolidate_supplier_casefold(df: pd.DataFrame) -> pd.DataFrame:
    """Map labels that differ only by case/whitespace/unicode form to one canonical string (longest wins)."""
    if df.empty or C_SUPPLIER_NAME not in df.columns:
        return df
    out = df.copy()
    s = out[C_SUPPLIER_NAME].fillna("Unknown").astype(str).str.strip()
    s = s.replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    buckets: dict[str, list[str]] = {}
    for val in pd.unique(s.values):
        kk = _normalize_supplier_group_key(val)
        if not kk:
            continue
        buckets.setdefault(kk, []).append(str(val))
    key_to_canon = {
        k: max(members, key=lambda x: (len(x), x))
        for k, members in buckets.items()
        if members
    }

    def pick(v: Any) -> str:
        if pd.isna(v):
            return "Unknown"
        vv = str(v).strip()
        if not vv or vv.lower() in ("nan", "none", ""):
            return "Unknown"
        kk = _normalize_supplier_group_key(vv)
        if not kk:
            return vv
        return key_to_canon.get(kk, vv)

    out[C_SUPPLIER_NAME] = s.map(pick)
    return out


def _canonicalize_supplier_names_by_match_key(ser: pd.Series) -> pd.Series:
    """Assign one display name per `_supplier_match_key` (longest label wins) so sums group correctly."""
    s = ser.fillna("Unknown").astype(str).str.strip()
    s = s.replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    buckets: dict[str, list[str]] = {}
    for u in pd.unique(s.values):
        uu = str(u).strip()
        kk = _supplier_match_key(uu)
        if not kk or uu.lower() in ("unknown", "nan", "none", ""):
            buckets.setdefault("__UNK__", []).append(uu)
            continue
        buckets.setdefault(kk, []).append(uu)
    key_to_canon: dict[str, str] = {}
    for k, members in buckets.items():
        if k == "__UNK__":
            key_to_canon[k] = "Unknown"
            continue
        key_to_canon[k] = _canonical_supplier_label_for_key(k, members)

    def pick(u: Any) -> str:
        if pd.isna(u):
            return "Unknown"
        uu = str(u).strip()
        if not uu or uu.lower() in ("unknown", "nan", "none", ""):
            return "Unknown"
        kk = _supplier_match_key(uu)
        if not kk:
            return uu
        return key_to_canon.get(kk, uu)

    return s.map(pick)


def _aggregate_by_supplier(df: pd.DataFrame) -> pd.DataFrame:
    """One row per supplier: summed amounts, merged sources / IDs / months."""
    if df.empty or C_SUPPLIER_NAME not in df.columns:
        return df

    work = df.copy()
    for col in (C_DEBIT, C_CREDIT, C_AMOUNT):
        if col in work.columns:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)

    work[C_SUPPLIER_NAME] = _canonicalize_supplier_names_by_match_key(work[C_SUPPLIER_NAME])

    group_cols = [C_SUPPLIER_NAME]

    agg_dict: dict = {}
    if C_DEBIT in work.columns:
        agg_dict[C_DEBIT] = "sum"
    if C_CREDIT in work.columns:
        agg_dict[C_CREDIT] = "sum"
    if C_AMOUNT in work.columns:
        agg_dict[C_AMOUNT] = "sum"

    if not agg_dict:
        return df

    extra_agg: dict = {}
    if C_SOURCE in work.columns:
        src = work[C_SOURCE].map(lambda v: str(v).strip() if pd.notna(v) else "")
        if (src != "").any():
            extra_agg[C_SOURCE] = _agg_join_unique_strings
    if C_SUPPLIER_ID in work.columns:
        extra_agg[C_SUPPLIER_ID] = _agg_join_unique_ids
    if C_GL_ACCOUNT in work.columns:
        gl = work[C_GL_ACCOUNT].map(lambda v: str(v).strip() if pd.notna(v) else "")
        if (gl != "").any():
            extra_agg[C_GL_ACCOUNT] = _agg_join_unique_strings
    if C_COST_CENTER in work.columns:
        cc = work[C_COST_CENTER].map(lambda v: str(v).strip() if pd.notna(v) else "")
        if (cc != "").any():
            extra_agg[C_COST_CENTER] = _agg_join_unique_strings
    if C_COMPANY in work.columns:
        co = work[C_COMPANY].map(lambda v: str(v).strip() if pd.notna(v) else "")
        if (co != "").any():
            extra_agg[C_COMPANY] = _agg_join_unique_strings
    if "Month" in work.columns:
        extra_agg["Month"] = _agg_join_months

    agg_dict.update(extra_agg)

    g = work.groupby(group_cols, as_index=False)
    agg_df = g.agg(agg_dict)
    agg_df.insert(1, "Transactions", work.groupby(group_cols).size().values)

    if C_DEBIT in agg_df.columns and C_CREDIT in agg_df.columns:
        agg_df["NetSpend"] = agg_df[C_DEBIT] - agg_df[C_CREDIT]
    elif C_AMOUNT in agg_df.columns:
        agg_df["NetSpend"] = agg_df[C_AMOUNT]

    first_cols = [C_SUPPLIER_NAME, "Transactions"]
    for c in (C_DEBIT, C_CREDIT, "NetSpend", C_AMOUNT):
        if c in agg_df.columns:
            first_cols.append(c)
    rest = [c for c in agg_df.columns if c not in first_cols]
    agg_df = agg_df[first_cols + rest]
    if C_DEBIT in agg_df.columns:
        sort_col = C_DEBIT
    elif C_AMOUNT in agg_df.columns:
        sort_col = C_AMOUNT
    elif C_CREDIT in agg_df.columns:
        sort_col = C_CREDIT
    else:
        sort_col = C_SUPPLIER_NAME
    agg_df = agg_df.sort_values(sort_col, ascending=False).reset_index(drop=True)
    dup_n = int(agg_df[C_SUPPLIER_NAME].duplicated().sum())
    if dup_n:
        logger.warning("_aggregate_by_supplier: unexpected duplicate supplier rows (%s)", dup_n)
    return agg_df


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="CleanedData")
    return buf.getvalue()


@st.cache_data(show_spinner=False, max_entries=64)
def _cached_process_file(upload_name: str, raw_bytes: bytes) -> dict:
    """Cache by filename + bytes so re-runs (widget interaction) stay fast."""

    class _Uploaded:
        def __init__(self, name: str, blob: bytes):
            self.name = name
            self._blob = blob

        def getvalue(self) -> bytes:
            return self._blob

    return process_file(_Uploaded(upload_name, raw_bytes))


EURO = "\u20ac"

_SESSION_KEYS = (
    "_erp_sig",
    "_erp_results",
    "_erp_all_tx",
    "_erp_all_suppliers",
    "_erp_all_pivots",
    "_erp_errors",
)

# Streamlit multiselect breaks or hangs with huge option lists (browser payload).
_MAX_SUPPLIER_MULTISELECT = 400


def _handle_file_error(filename: str, error: Exception, error_type: str = "unknown") -> dict:
    """Create user-friendly error message for file processing issues."""
    error_msg = str(error)

    if "encrypted" in error_msg.lower() or "password" in error_msg.lower():
        user_msg = f"❌ **{filename}**: File is password-protected. Please unlock it in Excel first and re-save."
        error_code = "PASSWORD_PROTECTED"
    elif "xlrd" in error_msg.lower() and filename.lower().endswith(".xls"):
        user_msg = f"❌ **{filename}**: Unsupported .xls format. Please convert to .xlsx first."
        error_code = "XLS_FORMAT"
    elif "truncated" in error_msg.lower() or "corrupt" in error_msg.lower():
        user_msg = f"❌ **{filename}**: File appears to be corrupted. Please re-download and try again."
        error_code = "CORRUPTED_FILE"
    elif "sheet" in error_msg.lower() and "not found" in error_msg.lower():
        user_msg = f"❌ **{filename}**: Expected sheets (Materialkonten/Kreditoren) not found."
        error_code = "MISSING_SHEETS"
    elif "encoding" in error_msg.lower() or "codec" in error_msg.lower():
        user_msg = f"❌ **{filename}**: File encoding issue. Try saving as UTF-8 in Excel."
        error_code = "ENCODING_ERROR"
    elif "memory" in error_msg.lower():
        user_msg = f"❌ **{filename}**: File is too large. Please split into smaller files."
        error_code = "OUT_OF_MEMORY"
    else:
        user_msg = f"❌ **{filename}**: {error_msg[:100]}{'...' if len(error_msg) > 100 else ''}"
        error_code = "UNKNOWN_ERROR"

    return {
        "filename": filename,
        "message": user_msg,
        "error_code": error_code,
        "technical_error": error_msg,
    }


def _show_error_modal(errors: list) -> None:
    """Display error messages in a user-friendly format."""
    if not errors:
        return

    with st.container():
        st.error("⚠️ **File Processing Errors**")
        for error in errors:
            st.markdown(error["message"])
            with st.expander("Technical details"):
                st.code(error["technical_error"], language="text")


def main() -> None:
    st.set_page_config(page_title="ERP Procurement Dashboard", layout="wide")
    st.markdown("""<style>
    [data-testid="stMetricValue"] { font-size: 1.15rem; }
    .block-container { padding-top: 1.5rem; }
    div[data-testid="stExpander"] details summary p { font-weight: 600; }
    </style>""", unsafe_allow_html=True)
    st.title("ERP Procurement Dashboard")

    with st.sidebar:
        st.header("Upload & Filters")
        uploaded_files = st.file_uploader("Upload Excel or CSV files", type=["xlsx", "xls", "csv"], accept_multiple_files=True)
        st.caption(
            "Uses **openpyxl** for .xlsx and **xlrd** for legacy .xls (see requirements.txt). "
            "Password-protected workbooks must be unlocked in Excel first. Max upload **500 MB**."
        )

    if not uploaded_files:
        for k in _SESSION_KEYS:
            st.session_state.pop(k, None)
        st.info("Upload one or more Excel/CSV files to get started.\n\n"
                "The dashboard **auto-detects** sheet types, column names, header rows, "
                "numeric fields, dates, and categories -- no manual configuration needed.")
        return

    fp = tuple(f.file_id for f in uploaded_files)

    if st.session_state.get("_erp_sig") != fp:
        blobs = [(f.name, f.getvalue()) for f in uploaded_files]
        results: list[dict] = []
        errors: list[dict] = []
        progress = st.progress(0, text="Processing files...")
        for i, (fname, raw) in enumerate(blobs):
            try:
                result = _cached_process_file(fname, raw)
                results.append(result)
                progress.progress((i + 1) / len(blobs), text=f"✓ Processed {fname}")
            except Exception as e:
                error_info = _handle_file_error(fname, e)
                errors.append(error_info)
                logger.error(f"File processing error: {fname} - {str(e)}")
                progress.progress((i + 1) / len(blobs), text=f"✗ Failed {fname}")
        progress.empty()

        # Show errors if any
        if errors:
            _show_error_modal(errors)
            st.session_state["_erp_errors"] = errors

        all_tx = (
            pd.concat(
                [r["transactions"] for r in results if not r["transactions"].empty],
                ignore_index=True,
                sort=False,
            )
            if any(not r["transactions"].empty for r in results)
            else pd.DataFrame()
        )
        all_suppliers = (
            pd.concat(
                [r["suppliers"] for r in results if not r["suppliers"].empty],
                ignore_index=True,
                sort=False,
            )
            if any(not r["suppliers"].empty for r in results)
            else pd.DataFrame()
        )
        all_pivots = (
            pd.concat(
                [r["pivots"] for r in results if not r["pivots"].empty],
                ignore_index=True,
                sort=False,
            )
            if any(not r["pivots"].empty for r in results)
            else pd.DataFrame()
        )

        if all_tx.empty and not all_pivots.empty:
            all_tx = all_pivots.copy()
            all_tx[C_DEBIT] = all_tx[C_AMOUNT].clip(lower=0)
            all_tx[C_CREDIT] = all_tx[C_AMOUNT].clip(upper=0).abs()
            if C_DATE not in all_tx.columns:
                all_tx[C_DATE] = pd.NaT
            if C_DESCRIPTION not in all_tx.columns:
                all_tx[C_DESCRIPTION] = ""

        if not all_tx.empty or not all_pivots.empty:
            if C_DATE in all_tx.columns and all_tx[C_DATE].notna().any():
                all_tx["Month"] = pd.to_datetime(all_tx[C_DATE], errors="coerce").dt.to_period("M").astype(str)
            else:
                all_tx["Month"] = "N/A"

            if C_SUPPLIER_NAME in all_tx.columns:
                sn = all_tx[C_SUPPLIER_NAME].fillna("Unknown").astype(str).str.strip()
                sn = sn.replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
                all_tx[C_SUPPLIER_NAME] = sn
                n_distinct = all_tx[C_SUPPLIER_NAME].nunique()
                n_rows_tx = len(all_tx)
                if n_distinct > 60_000:
                    logger.warning(
                        "Skipping supplier clustering (%s distinct names on %s rows).",
                        n_distinct,
                        n_rows_tx,
                    )
                else:
                    non_unknown = sorted(
                        {
                            x
                            for x in all_tx[C_SUPPLIER_NAME].unique().tolist()
                            if str(x).strip().lower() not in ("unknown", "nan", "")
                        }
                    )
                    if len(non_unknown) > 25_000:
                        logger.info(
                            "Skipping match-key supplier collapse (%s unique labels).",
                            len(non_unknown),
                        )
                    elif len(non_unknown) > 1:
                        mk = _collapse_names_by_match_key(non_unknown)
                        all_tx[C_SUPPLIER_NAME] = _series_apply_name_map(all_tx[C_SUPPLIER_NAME], mk)
                    u2 = [
                        x
                        for x in all_tx[C_SUPPLIER_NAME].dropna().unique().tolist()
                        if str(x).strip().lower() not in ("unknown", "nan", "")
                    ]
                    if 1 < len(u2) < 2_500 and n_rows_tx < 800_000:
                        name_map = _fuzzy_group_names(u2, threshold=FUZZY_SUPPLIER_MERGE_THRESHOLD)
                        all_tx[C_SUPPLIER_NAME] = _series_apply_name_map(all_tx[C_SUPPLIER_NAME], name_map)

            if not all_suppliers.empty and C_SUPPLIER_ID in all_suppliers.columns:
                sup_dedup = all_suppliers.copy()
                sup_dedup["_nlen"] = sup_dedup[C_SUPPLIER_NAME].astype(str).str.len()
                sup_dedup = sup_dedup.sort_values("_nlen", ascending=False).drop_duplicates(subset=[C_SUPPLIER_ID]).drop(
                    columns=["_nlen"]
                )
                all_suppliers = sup_dedup.reset_index(drop=True)

        st.session_state["_erp_sig"] = fp
        st.session_state["_erp_results"] = results
        st.session_state["_erp_all_tx"] = all_tx
        st.session_state["_erp_all_suppliers"] = all_suppliers
        st.session_state["_erp_all_pivots"] = all_pivots

    try:
        results = st.session_state["_erp_results"]
        all_tx = st.session_state["_erp_all_tx"]
        all_suppliers = st.session_state["_erp_all_suppliers"]
        all_pivots = st.session_state["_erp_all_pivots"]
    except KeyError:
        st.session_state.pop("_erp_sig", None)
        st.rerun()

    if isinstance(all_tx, pd.DataFrame):
        all_tx = all_tx.copy()
        if "Month" not in all_tx.columns:
            if C_DATE in all_tx.columns and all_tx[C_DATE].notna().any():
                all_tx["Month"] = (
                    pd.to_datetime(all_tx[C_DATE], errors="coerce").dt.to_period("M").astype(str)
                )
            else:
                all_tx["Month"] = "N/A"
        if C_SUPPLIER_NAME in all_tx.columns and not all_tx.empty:
            all_tx = _consolidate_supplier_casefold(all_tx)
            all_tx[C_SUPPLIER_NAME] = _canonicalize_supplier_names_by_match_key(all_tx[C_SUPPLIER_NAME])

    if all_tx.empty and all_pivots.empty:
        # Show any processing errors first
        if st.session_state.get("_erp_errors"):
            _show_error_modal(st.session_state["_erp_errors"])

        st.error("❌ No transaction data could be extracted from the uploaded files.")
        st.info(
            "**Possible causes:**\n"
            "- Sheets are empty or contain only summaries\n"
            "- Header row was not detected correctly\n"
            "- File format or encoding issue\n"
            "- Password-protected or corrupted file\n\n"
            "Open **File Processing Details** below to see diagnostics."
        )
        with st.expander("File Processing Details", expanded=True):
            for r in results:
                st.markdown(f"**{r['filename']}**")
                if r["sheet_info"]:
                    st.dataframe(pd.DataFrame(r["sheet_info"]), use_container_width=True, hide_index=True)
                else:
                    st.caption("No sheet metadata recorded.")
        return

    nf = len(results)
    st.success(
        f"Loaded **{len(all_tx):,}** transaction rows from **{nf}** file{'s' if nf != 1 else ''}. "
        f"Similar supplier names are merged at **{int(FUZZY_SUPPLIER_MERGE_THRESHOLD * 100)}%** match; "
        f"amounts sum to **one row per supplier** in the summary below. Use sidebar filters as needed."
    )

    with st.expander("Uploaded Files Overview", expanded=False):
        for r in results:
            st.markdown(f"**{r['filename']}** -- Company: *{r['company']}*")
            st.dataframe(pd.DataFrame(r["sheet_info"]), use_container_width=True, hide_index=True)

    with st.sidebar:
        st.markdown("---")
        selected_companies = []
        if C_SOURCE in all_tx.columns:
            selected_companies = st.multiselect(
                "Company / File",
                options=_safe_sort_filter_options(all_tx[C_SOURCE].dropna().unique().tolist()),
                default=[],
            )
        selected_suppliers = []
        if C_SUPPLIER_NAME in all_tx.columns:
            supplier_all = sorted(
                {
                    str(s).strip()
                    for s in all_tx[C_SUPPLIER_NAME].dropna().tolist()
                    if str(s).strip() and str(s).strip().lower() not in ("nan", "unknown")
                }
            )
            if not supplier_all:
                selected_suppliers = []
            elif len(supplier_all) > _MAX_SUPPLIER_MULTISELECT:
                st.caption(
                    f"{len(supplier_all):,} suppliers — use **Supplier name contains** to narrow "
                    f"(multiselect shows up to {_MAX_SUPPLIER_MULTISELECT} matches)."
                )
                q = st.text_input("Supplier name contains", value="", key="erp_supplier_search")
                qlow = q.strip().lower()
                if qlow:
                    filtered = [x for x in supplier_all if qlow in x.lower()][: _MAX_SUPPLIER_MULTISELECT]
                else:
                    filtered = supplier_all[: _MAX_SUPPLIER_MULTISELECT]
                selected_suppliers = st.multiselect("Supplier", options=filtered, default=[])
            else:
                selected_suppliers = st.multiselect("Supplier", options=supplier_all, default=[])
        month_options = (
            _safe_sort_filter_options([m for m in all_tx["Month"].dropna().unique().tolist() if m != "N/A"])
            if "Month" in all_tx.columns
            else []
        )
        selected_months = st.multiselect("Month", options=month_options, default=[])
        extra_filters: dict[str, list] = {}
        for col in [c for c in [C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY] if c in all_tx.columns]:
            vals = all_tx[col].dropna()
            vals = vals[vals.astype(str).str.strip() != ""]
            uv = _safe_sort_filter_options(vals.unique().tolist())
            if 2 <= len(uv) <= 100:
                sel = st.multiselect(col, options=uv, default=[])
                if sel:
                    extra_filters[col] = sel

    view = all_tx.copy()
    if selected_companies:
        view = view[view[C_SOURCE].isin(selected_companies)]
    if selected_suppliers:
        view = view[view[C_SUPPLIER_NAME].isin(selected_suppliers)]
    if selected_months:
        view = view[view["Month"].isin(selected_months)]
    for col, vals in extra_filters.items():
        view = view[view[col].isin(vals)]

    total_debit = view[C_DEBIT].sum() if C_DEBIT in view.columns else 0
    total_credit = view[C_CREDIT].sum() if C_CREDIT in view.columns else 0
    total_net = view[C_AMOUNT].sum() if C_AMOUNT in view.columns else 0
    n_suppliers = view[C_SUPPLIER_NAME].nunique() if C_SUPPLIER_NAME in view.columns else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Total Debit (Spend)", f"{EURO} {total_debit:,.2f}")
    k2.metric("Total Credit (Returns)", f"{EURO} {total_credit:,.2f}")
    k3.metric("Net Amount", f"{EURO} {total_net:,.2f}")
    k4.metric("Unique Suppliers", f"{n_suppliers:,}")
    k5.metric("Files / Records", f"{len(uploaded_files)} / {len(view):,}")
    st.markdown("---")

    agg_view = _aggregate_by_supplier(view)
    st.subheader("Supplier summary — one row per supplier")
    st.caption(
        f"Names that are **≥{int(FUZZY_SUPPLIER_MERGE_THRESHOLD * 100)}%** similar (fuzzy match) are treated as one supplier; "
        "**Debit**, **Credit**, and **Amount** are **added** across all matching rows. "
        "Transaction count, months, and sources are combined. Detail lines stay in 'All Individual Transactions'."
    )
    st.dataframe(agg_view, use_container_width=True, height=420, hide_index=True)
    st.markdown("---")

    col_left, col_right = st.columns(2)
    spend_view = view[view[C_DEBIT] > 0] if C_DEBIT in view.columns else view

    with col_left:
        st.subheader("Top 20 Suppliers by Spend")
        if C_SUPPLIER_NAME in spend_view.columns and not spend_view.empty:
            top_sup = spend_view.groupby(C_SUPPLIER_NAME, as_index=False)[C_DEBIT].sum().sort_values(C_DEBIT, ascending=False).head(20)
            if not top_sup.empty:
                fig = px.bar(top_sup, x=C_SUPPLIER_NAME, y=C_DEBIT, color=C_DEBIT,
                             color_continuous_scale="Blues", labels={C_DEBIT: f"Spend ({EURO})", C_SUPPLIER_NAME: "Supplier"})
                fig.update_layout(xaxis_tickangle=-45, showlegend=False, coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No spend data.")
        else:
            st.info("No spend data.")

    with col_right:
        st.subheader("Monthly Spend Trend")
        vm = spend_view[spend_view["Month"] != "N/A"] if "Month" in spend_view.columns else pd.DataFrame()
        if not vm.empty and C_DEBIT in vm.columns:
            monthly = vm.groupby("Month", as_index=False)[C_DEBIT].sum().sort_values("Month")
            fig = px.line(monthly, x="Month", y=C_DEBIT, markers=True,
                          labels={C_DEBIT: f"Spend ({EURO})", "Month": "Month"})
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Date information not available for monthly chart.")

    st.markdown("---")

    if C_SOURCE in view.columns and view[C_SOURCE].nunique() > 1:
        st.subheader("Spend by Company / File")
        c1, c2 = st.columns(2)
        by_company = view.groupby(C_SOURCE, as_index=False)[C_DEBIT].sum().sort_values(C_DEBIT, ascending=False)
        with c1:
            fig = px.bar(by_company, x=C_SOURCE, y=C_DEBIT, color=C_DEBIT,
                         color_continuous_scale="Greens", labels={C_DEBIT: f"Spend ({EURO})", C_SOURCE: "Company"})
            fig.update_layout(xaxis_tickangle=-45, showlegend=False, coloraxis_showscale=False)
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            fig = px.pie(by_company, names=C_SOURCE, values=C_DEBIT, labels={C_DEBIT: f"Spend ({EURO})", C_SOURCE: "Company"})
            fig.update_traces(textposition="inside", textinfo="percent+label")
            st.plotly_chart(fig, use_container_width=True)
        st.markdown("---")

    if C_SOURCE in view.columns and view[C_SOURCE].nunique() > 1 and "Month" in view.columns:
        vm2 = view[view["Month"] != "N/A"]
        if not vm2.empty:
            st.subheader("Monthly Spend by Company")
            mc = vm2.groupby([C_SOURCE, "Month"], as_index=False)[C_DEBIT].sum().sort_values("Month")
            fig = px.line(mc, x="Month", y=C_DEBIT, color=C_SOURCE, markers=True,
                          labels={C_DEBIT: f"Spend ({EURO})", "Month": "Month", C_SOURCE: "Company"})
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
            st.markdown("---")

    breakdown_cols = [c for c in [C_GL_ACCOUNT, C_COST_CENTER, C_CATEGORY]
                      if c in view.columns and view[c].astype(str).str.strip().replace("", pd.NA).dropna().nunique() >= 2]
    if breakdown_cols:
        st.subheader("Spend Breakdown")
        tabs = st.tabs(breakdown_cols)
        for tab, col in zip(tabs, breakdown_cols):
            with tab:
                bd = view[view[col].astype(str).str.strip() != ""].groupby(col, as_index=False)[C_DEBIT].sum().sort_values(C_DEBIT, ascending=False).head(25)
                if not bd.empty:
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        fig = px.bar(bd, x=col, y=C_DEBIT, color=C_DEBIT, color_continuous_scale="Oranges", labels={C_DEBIT: f"Spend ({EURO})"})
                        fig.update_layout(xaxis_tickangle=-45, showlegend=False, coloraxis_showscale=False)
                        st.plotly_chart(fig, use_container_width=True)
                    with bc2:
                        fig = px.pie(bd.head(10), names=col, values=C_DEBIT)
                        fig.update_traces(textposition="inside", textinfo="percent+label")
                        st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info(f"No data for {col}.")
        st.markdown("---")

    if C_SUPPLIER_NAME in view.columns and not view.empty:
        st.subheader("Supplier Debit vs Credit")
        sc = view.groupby(C_SUPPLIER_NAME, as_index=False).agg({C_DEBIT: "sum", C_CREDIT: "sum"})
        sc = sc[(sc[C_DEBIT] > 0) | (sc[C_CREDIT] > 0)]
        if not sc.empty and len(sc) > 1:
            sc["_size"] = (sc[C_DEBIT].abs() + sc[C_CREDIT].abs()).clip(lower=1)
            fig = px.scatter(sc, x=C_DEBIT, y=C_CREDIT, hover_name=C_SUPPLIER_NAME,
                             size="_size",
                             labels={C_DEBIT: f"Total Debit ({EURO})", C_CREDIT: f"Total Credit ({EURO})"}, size_max=40)
            st.plotly_chart(fig, use_container_width=True)
            st.markdown("---")

    if C_AMOUNT in view.columns and len(view) > 10:
        st.subheader("Transaction Amount Distribution")
        amounts = view[C_AMOUNT][view[C_AMOUNT] != 0]
        if len(amounts) > 5:
            fig = px.histogram(amounts, nbins=50, labels={"value": f"Amount ({EURO})", "count": "Frequency"},
                               color_discrete_sequence=["#636EFA"])
            fig.update_layout(showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
            st.markdown("---")

    vm3 = view[view["Month"] != "N/A"] if "Month" in view.columns else pd.DataFrame()
    if not vm3.empty and C_SUPPLIER_NAME in vm3.columns and vm3["Month"].nunique() >= 2:
        st.subheader("Top 15 Suppliers -- Monthly Heatmap")
        top15 = vm3.groupby(C_SUPPLIER_NAME)[C_DEBIT].sum().nlargest(15).index.tolist()
        hm = vm3[vm3[C_SUPPLIER_NAME].isin(top15)].groupby([C_SUPPLIER_NAME, "Month"])[C_DEBIT].sum().unstack(fill_value=0)
        if not hm.empty:
            fig = px.imshow(hm, labels=dict(x="Month", y="Supplier", color=f"Spend ({EURO})"),
                            color_continuous_scale="YlOrRd", aspect="auto")
            fig.update_layout(height=500)
            st.plotly_chart(fig, use_container_width=True)
            st.markdown("---")

    with st.expander("All Individual Transactions", expanded=False):
        st.dataframe(view, use_container_width=True, height=400, hide_index=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button("Download Supplier Summary (Excel)", data=to_excel_bytes(agg_view),
                           file_name="supplier_summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c2:
        st.download_button("Download All Transactions (Excel)", data=to_excel_bytes(view),
                           file_name="cleaned_procurement_data.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with c3:
        if not all_suppliers.empty:
            st.download_button("Download Suppliers (Excel)", data=to_excel_bytes(all_suppliers),
                               file_name="cleaned_suppliers.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    main()
